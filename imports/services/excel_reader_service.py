from datetime import datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from .import_validation_service import (
    EXPECTED_SHEET_NAME,
    validate_headers,
    validate_sheet_name,
)


MAX_EMPTY_ROWS = 200
HEADER_ROW = 1

FIELD_BY_POSITION = {
    1: "map_date",
    2: "map_number",
    3: "owner",
    4: "route",
    5: "delivery_points",
    6: "weight",
    7: "volume",
    8: "map_value",
    9: "vehicle_plate",
    10: "driver_code",
    11: "driver_name",
    12: "checker_code",
    13: "checker_name",
    14: "load_status_description",
    15: "load_date",
    16: "invoice_number",
    17: "invoice_series",
    18: "invoice_issue_date",
    19: "bordero_number",
    20: "customer_code",
    21: "customer_name",
    22: "city",
    23: "invoice_value",
    25: "invoice_status",
    26: "cargo_status",
    27: "auxiliary_date",
    28: "delivery_status",
    29: "customer_delivery_date",
    30: "customer_delivery_time",
    31: "exported_to",
    32: "notes",
    33: "seller_code",
    34: "team_code",
}

REQUIRED_FIELDS = {
    "map_date",
    "map_number",
    "owner",
    "route",
    "delivery_points",
    "weight",
    "volume",
    "map_value",
    "vehicle_plate",
    "load_status_description",
    "load_date",
    "invoice_number",
    "invoice_series",
    "invoice_issue_date",
    "bordero_number",
    "customer_code",
    "customer_name",
    "city",
    "invoice_value",
    "invoice_status",
    "cargo_status",
    "delivery_status",
    "seller_code",
    "team_code",
}


class RowNormalizationError(Exception):
    pass


def read_lead_time_rows(file_path):
    workbook = openpyxl.load_workbook(
        Path(file_path),
        read_only=True,
        data_only=True,
    )
    try:
        validate_sheet_name(workbook)

        sheet = workbook[EXPECTED_SHEET_NAME]
        header_values = next(
            sheet.iter_rows(
                min_row=HEADER_ROW,
                max_row=HEADER_ROW,
                values_only=True,
            )
        )
        validate_headers(header_values)

        empty_rows = 0
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if _is_empty_row(row):
                empty_rows += 1
                if empty_rows >= MAX_EMPTY_ROWS:
                    break
                continue

            empty_rows = 0
            try:
                yield row_number, _normalize_row(row), None
            except RowNormalizationError as exc:
                yield row_number, None, exc
    finally:
        workbook.close()


def _normalize_row(row):
    normalized = {}
    for position, field_name in FIELD_BY_POSITION.items():
        value = row[position - 1] if len(row) >= position else None
        normalized[field_name] = _clean_value(value)

    _coerce_types(normalized)
    _validate_required_fields(normalized)
    return normalized


def _coerce_types(row):
    date_fields = [
        "map_date",
        "load_date",
        "invoice_issue_date",
        "auxiliary_date",
        "customer_delivery_date",
    ]
    decimal_fields = ["weight", "volume", "map_value", "invoice_value"]
    integer_fields = ["delivery_points"]
    text_fields = [
        "map_number",
        "owner",
        "route",
        "vehicle_plate",
        "driver_code",
        "driver_name",
        "checker_code",
        "checker_name",
        "load_status_description",
        "invoice_number",
        "invoice_series",
        "bordero_number",
        "customer_code",
        "customer_name",
        "city",
        "invoice_status",
        "cargo_status",
        "delivery_status",
        "exported_to",
        "notes",
        "seller_code",
        "team_code",
    ]

    for field_name in date_fields:
        row[field_name] = _to_date(row[field_name], field_name)
    for field_name in decimal_fields:
        row[field_name] = _to_decimal(row[field_name], field_name)
    for field_name in integer_fields:
        row[field_name] = _to_int(row[field_name], field_name)
    for field_name in text_fields:
        row[field_name] = _to_text(row[field_name])

    row["customer_delivery_time"] = _to_time(
        row["customer_delivery_time"],
        "customer_delivery_time",
    )


def _validate_required_fields(row):
    missing_fields = [
        field_name
        for field_name in REQUIRED_FIELDS
        if row.get(field_name) in (None, "")
    ]
    if missing_fields:
        raise RowNormalizationError(
            f"Campos obrigatórios ausentes: {', '.join(sorted(missing_fields))}"
        )

    if row["delivery_status"].strip().upper() == "ENTREGUE":
        if not row["customer_delivery_date"] or not row["customer_delivery_time"]:
            raise RowNormalizationError(
                "Entrega marcada como ENTREGUE sem data ou hora de entrega."
            )


def _clean_value(value):
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _is_empty_row(row):
    return not any(_clean_value(value) is not None for value in row)


def _to_text(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def _to_date(value, field_name):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return value
    if isinstance(value, str):
        for date_format in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), date_format).date()
            except ValueError:
                continue
    raise RowNormalizationError(f"Data inválida em {field_name}: {value}")


def _to_time(value, field_name):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, str):
        for time_format in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(value.strip(), time_format).time()
            except ValueError:
                continue
    raise RowNormalizationError(f"Hora inválida em {field_name}: {value}")


def _to_decimal(value, field_name):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError) as exc:
        raise RowNormalizationError(f"Número inválido em {field_name}: {value}") from exc


def _to_int(value, field_name):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise RowNormalizationError(f"Inteiro inválido em {field_name}: {value}") from exc
