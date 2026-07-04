import json
from io import BytesIO
from datetime import date, time
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.http import QueryDict
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
import openpyxl

from accounts.services.rbac_service import LOGISTICA_ADMIN, LOGISTICA_VIEWER, setup_rbac
from dashboard.services.dashboard_filters import DashboardFilters
from dashboard.services.dashboard_queries import get_dashboard_context
from dashboard.services.export_service import XLSX_CONTENT_TYPE
from imports.models import ImportBatch, LeadTimeRecord


User = get_user_model()


class DashboardViewTests(TestCase):
    def setUp(self):
        setup_rbac()
        self.user = User.objects.create_user(
            username="plain",
            password="safe-test-password",
        )
        self.viewer = User.objects.create_user(
            username="viewer",
            password="safe-test-password",
        )
        self.admin = User.objects.create_user(
            username="admin",
            password="safe-test-password",
        )
        self.viewer.groups.add(Group.objects.get(name=LOGISTICA_VIEWER))
        self.admin.groups.add(Group.objects.get(name=LOGISTICA_ADMIN))

    def test_dashboard_requires_login(self):
        response = self.client.get(reverse("dashboard:home"))

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("accounts:login"), response["Location"])

    def test_authenticated_user_without_dashboard_permission_is_blocked(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("dashboard:home"))

        self.assertEqual(response.status_code, 403)

    def test_logistica_viewer_can_access_dashboard(self):
        self.client.force_login(self.viewer)

        response = self.client.get(reverse("dashboard:home"))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "dashboard/home.html")

    def test_logistica_admin_can_access_dashboard(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse("dashboard:home"))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "dashboard/home.html")

    def test_dashboard_get_filters_do_not_break_view(self):
        self.client.force_login(self.viewer)

        response = self.client.get(
            reverse("dashboard:home"),
            {
                "date_start": "2026-05-01",
                "date_end": "2026-05-31",
                "driver_name": "BEATRIZ GOMES",
                "route": "RT030",
                "business_unit": "TABACO",
                "region": "SUL",
                "frequency": "DIARIA",
                "delivery_status": "ENTREGUE",
                "cargo_status": "Ativa",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["filters"]["date_start"], "2026-05-01")
        self.assertEqual(response.context["filters"]["route"], ["RT030"])
        self.assertEqual(response.context["filters"]["region"], ["SUL"])
        self.assertEqual(response.context["filters"]["frequency"], ["DIARIA"])

    def test_dashboard_context_contains_expected_contract_keys(self):
        self.client.force_login(self.viewer)

        response = self.client.get(reverse("dashboard:home"))

        self.assertIn("cards", response.context)
        self.assertIn("charts", response.context)
        self.assertIn("filter_options", response.context)
        self.assertIn("metadata", response.context)
        self.assertIn("explanations", response.context)
        self.assertIn("table_pagination", response.context)

    def test_dashboard_renders_without_data(self):
        self.client.force_login(self.viewer)

        response = self.client.get(reverse("dashboard:home"))

        self.assertContains(response, "Nenhum registro encontrado")
        self.assertFalse(response.context["metadata"]["has_data"])

    def test_dashboard_renders_chart_json_and_canvas_ids(self):
        self.client.force_login(self.viewer)

        response = self.client.get(reverse("dashboard:home"))

        self.assertContains(response, 'id="dashboard-charts-data"')
        self.assertContains(response, 'id="chart-records-by-day"')
        self.assertContains(response, 'id="chart-driver-efficiency-scatter"')
        self.assertContains(response, 'id="chart-critical-routes-ranking"')
        self.assertContains(response, 'id="chart-weekday-bottleneck"')
        self.assertContains(response, 'id="chart-delay-pareto"')
        self.assertContains(response, 'id="chart-lead-time-distribution"')
        self.assertContains(response, 'id="chart-region-lead-time-comparison"')
        self.assertContains(response, 'id="chart-frequency-lead-time-comparison"')
        self.assertContains(response, 'id="chart-billing-vs-delivery-by-day"')
        self.assertContains(response, 'id="chart-delay-by-issue-day"')
        self.assertContains(response, 'class="chart-frame"', count=10)
        self.assertContains(response, "dashboard_charts.js")
        self.assertContains(response, "dashboard_help.js")
        self.assertContains(response, 'id="dashboard-explanations-data"')
        self.assertContains(response, 'data-help-key="cards.operational_sla_rate"')
        self.assertContains(response, 'data-help-key="charts.region_lead_time_comparison"')
        self.assertContains(response, "Pressão comercial")
        self.assertContains(response, 'data-help-key="charts.billing_vs_delivery_by_day"')
        self.assertContains(response, "dashboard_filters.js")

    def test_home_links_to_dashboard(self):
        self.client.force_login(self.viewer)

        response = self.client.get(reverse("core:home"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("dashboard:home"))


class DashboardExportTests(TestCase):
    def setUp(self):
        setup_rbac()
        self.user = User.objects.create_user(
            username="plain",
            password="safe-test-password",
        )
        self.access_only_user = User.objects.create_user(
            username="access-only",
            password="safe-test-password",
        )
        self.viewer = User.objects.create_user(
            username="viewer",
            password="safe-test-password",
        )
        self.admin = User.objects.create_user(
            username="admin",
            password="safe-test-password",
        )
        self.access_only_user.user_permissions.add(Permission.objects.get(codename="access_dashboard"))
        self.viewer.groups.add(Group.objects.get(name=LOGISTICA_VIEWER))
        self.admin.groups.add(Group.objects.get(name=LOGISTICA_ADMIN))
        self.export_url = reverse("dashboard:export_excel")

    def test_anonymous_user_cannot_access_export(self):
        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("accounts:login"), response["Location"])

    def test_authenticated_user_without_export_permission_cannot_access_export(self):
        self.client.force_login(self.user)

        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 403)

    def test_user_with_must_change_password_is_redirected_before_exporting(self):
        self.viewer.profile.must_change_password = True
        self.viewer.profile.save(update_fields=["must_change_password"])
        self.client.force_login(self.viewer)

        response = self.client.get(self.export_url)

        self.assertRedirects(response, reverse("accounts:force_password_change"))

    def test_logistica_viewer_can_export_excel(self):
        self.client.force_login(self.viewer)

        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 200)

    def test_logistica_admin_can_export_excel(self):
        self.client.force_login(self.admin)

        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 200)

    def test_export_response_has_xlsx_headers(self):
        self.client.force_login(self.viewer)

        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("dashboard_lead_time_", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_export_generates_valid_workbook_without_data(self):
        self.client.force_login(self.viewer)

        response = self.client.get(self.export_url)
        workbook = openpyxl.load_workbook(BytesIO(response.content))

        self.assertIn("Dados filtrados", workbook.sheetnames)
        worksheet = workbook["Dados filtrados"]
        self.assertEqual(worksheet.max_row, 1)
        self.assertEqual(worksheet["A1"].value, "Unidade de negócio")
        self.assertEqual(worksheet["G1"].value, "Região")
        self.assertEqual(worksheet["H1"].value, "Frequência")

    def test_export_with_filters_returns_only_filtered_records(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            driver_name="BEATRIZ GOMES",
            route="RT030",
            business_unit="TABACO",
            region="SUL",
            frequency="DIARIA",
            delivery_status="ENTREGUE",
            cargo_status="Ativa",
            invoice_issue_date=date(2026, 5, 1),
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            driver_name="ANA SILVA",
            route="RT010",
            business_unit="ALIMENTOS",
            delivery_status="PENDENTE",
            cargo_status="Cancelada",
            invoice_issue_date=date(2026, 6, 1),
        )
        self.client.force_login(self.viewer)

        response = self.client.get(
            self.export_url,
            {
                "date_start": "2026-05-01",
                "date_end": "2026-05-31",
                "driver_name": "BEATRIZ GOMES",
                "route": "RT030",
                "business_unit": "TABACO",
                "region": "SUL",
                "frequency": "DIARIA",
                "delivery_status": "ENTREGUE",
                "cargo_status": "Ativa",
            },
        )
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        worksheet = workbook["Dados filtrados"]

        self.assertEqual(worksheet.max_row, 2)
        self.assertEqual(worksheet["G2"].value, "SUL")
        self.assertEqual(worksheet["H2"].value, "DIARIA")
        self.assertEqual(worksheet["J2"].value, "1001")
        self.assertEqual(worksheet["D2"].value, "BEATRIZ GOMES")
        self.assertEqual(worksheet["E2"].value, "RT030")

    def test_export_with_multi_select_filters_returns_matching_records(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            driver_name="BEATRIZ GOMES",
            route="RT030",
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            driver_name="ANA SILVA",
            route="RT010",
        )
        self._create_record(
            batch=batch,
            row_number=3,
            invoice_number="1003",
            driver_name="CARLOS LIMA",
            route="RT099",
        )
        query = QueryDict("", mutable=True)
        query.setlist("driver_name", ["BEATRIZ GOMES", "ANA SILVA"])
        query.setlist("route", ["RT030", "RT010"])
        self.client.force_login(self.viewer)

        response = self.client.get(self.export_url, query)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        worksheet = workbook["Dados filtrados"]

        self.assertEqual(worksheet.max_row, 3)
        self.assertEqual({worksheet["J2"].value, worksheet["J3"].value}, {"1001", "1002"})

    def test_table_export_returns_selected_dashboard_table(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            route="RT030",
            delivery_status="ENTREGUE",
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            route="RT010",
        )
        self.client.force_login(self.viewer)

        response = self.client.get(
            self.export_url,
            {
                "table": "critical_routes",
                "route": "RT030",
            },
        )
        self.assertEqual(response.status_code, 200)

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        worksheet = workbook["Rotas criticas"]

        self.assertIn("dashboard_lead_time_rotas_criticas_", response["Content-Disposition"])
        self.assertEqual(workbook.sheetnames, ["Rotas criticas"])
        self.assertEqual(worksheet["A1"].value, "Pauta/Rota")
        self.assertEqual(worksheet["A2"].value, "RT030")
        self.assertEqual(worksheet["B2"].value, 1)

    def test_export_button_is_visible_for_user_with_permission(self):
        self.client.force_login(self.viewer)

        response = self.client.get(reverse("dashboard:home"), {"route": "RT030"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Exportar Excel")
        self.assertContains(response, f"{self.export_url}?route=RT030")
        self.assertContains(response, f"{self.export_url}?route=RT030&table=driver_outliers")
        self.assertContains(response, f"{self.export_url}?route=RT030&table=critical_routes")

    def test_export_button_is_hidden_for_user_without_export_permission(self):
        self.client.force_login(self.access_only_user)

        response = self.client.get(reverse("dashboard:home"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Exportar Excel")

    def _create_batch(self, status=ImportBatch.Status.SUCCESS, file_hash="export-file-hash"):
        now = timezone.now()
        return ImportBatch.objects.create(
            source_mode=ImportBatch.SourceMode.LOCAL,
            file_name=f"{file_hash}.xlsx",
            file_path="data/test.xlsx",
            file_hash=file_hash,
            sheet_name="COM 001",
            status=status,
            started_at=now,
            finished_at=now if status == ImportBatch.Status.SUCCESS else None,
            total_rows=1,
            valid_rows=1 if status == ImportBatch.Status.SUCCESS else 0,
            invalid_rows=0,
        )

    def _create_record(self, batch, row_number, invoice_number, **overrides):
        defaults = {
            "row_hash": f"export-row-hash-{row_number}",
            "business_unit": "TABACO",
            "map_date": date(2026, 5, 1),
            "map_number": "470939",
            "owner": "948 TAINARA CAMARGO MONTE",
            "route": "RT030",
            "delivery_points": 18,
            "weight": Decimal("151.670"),
            "volume": Decimal("0.640"),
            "map_value": Decimal("375.10"),
            "vehicle_plate": "AZP8G94",
            "driver_code": "395",
            "driver_name": "BEATRIZ GOMES",
            "checker_code": "",
            "checker_name": "",
            "load_status_description": "CARREGADO EM",
            "load_date": date(2026, 5, 5),
            "invoice_series": "11",
            "invoice_issue_date": date(2026, 5, 4),
            "bordero_number": "807679",
            "customer_code": "6248888",
            "customer_name": "CLIENTE TESTE LTDA",
            "city": "PONTA GROSSA",
            "region": "",
            "frequency": "",
            "invoice_value": Decimal("1674.40"),
            "invoice_status": "AZP8G94",
            "cargo_status": "Ativa",
            "auxiliary_date": None,
            "delivery_status": "ENTREGUE",
            "customer_delivery_date": date(2026, 5, 5),
            "customer_delivery_time": time(16, 12),
            "customer_delivery_datetime": timezone.now(),
            "exported_to": "UMOV",
            "notes": "",
            "seller_code": "681",
            "team_code": "1",
            "operational_lead_time_hours": Decimal("10.00"),
            "carrier_lead_time_hours": Decimal("5.00"),
            "is_operational_late": False,
            "is_carrier_late": False,
            "business_days_count": 1,
        }
        defaults.update(overrides)

        return LeadTimeRecord.objects.create(
            import_batch=batch,
            row_number=row_number,
            invoice_number=invoice_number,
            **defaults,
        )


class DashboardAnalyticsTests(TestCase):
    def test_empty_dashboard_context_is_serializable(self):
        context = get_dashboard_context(DashboardFilters())

        self.assertEqual(context["cards"]["total_records"], 0)
        self.assertEqual(context["cards"]["total_invoice_value"], "0.00")
        self.assertEqual(context["cards"]["operational_sla_rate"], "0.00")
        self.assertEqual(context["cards"]["carrier_sla_rate"], "0.00")
        self.assertEqual(context["cards"]["operational_lead_time_p90_hours"], "0.00")
        self.assertEqual(context["cards"]["carrier_lead_time_p90_hours"], "0.00")
        self.assertEqual(context["cards"]["delayed_invoice_value"], "0.00")
        self.assertEqual(context["cards"]["status_inconsistency_count"], 0)
        self.assertEqual(context["cards"]["status_inconsistency_percentage"], "0.00")
        self.assertEqual(context["cards"]["peak_billing_day"]["records"], 0)
        self.assertEqual(context["cards"]["last_3_business_days_records"], 0)
        self.assertEqual(context["cards"]["last_3_business_days_percentage"], "0.00")
        self.assertEqual(context["cards"]["normal_daily_average_records"], "0.00")
        self.assertEqual(context["cards"]["top_critical_route"]["route"], "Sem dados")
        self.assertFalse(context["metadata"]["has_data"])
        self.assertIsNone(context["metadata"]["last_successful_import"])
        self.assertEqual(context["metadata"]["operational_target_hours"], "48.00")
        self.assertEqual(context["metadata"]["carrier_target_hours"], "24.00")
        self.assertIn("records_by_day", context["charts"])
        self.assertIn("driver_efficiency_scatter", context["charts"])
        self.assertIn("critical_routes_ranking", context["charts"])
        self.assertIn("weekday_bottleneck", context["charts"])
        self.assertIn("delay_pareto", context["charts"])
        self.assertIn("lead_time_distribution", context["charts"])
        self.assertIn("region_lead_time_comparison", context["charts"])
        self.assertIn("frequency_lead_time_comparison", context["charts"])
        self.assertIn("billing_vs_delivery_by_day", context["charts"])
        self.assertIn("delay_by_issue_day", context["charts"])
        self.assertIn("driver_outliers", context["tables"])
        self.assertIn("critical_routes", context["tables"])
        self.assertIn("critical_cities", context["tables"])
        self.assertIn("critical_regions", context["tables"])
        self.assertIn("critical_frequencies", context["tables"])
        self.assertIn("invoice_outliers", context["tables"])
        self.assertIn("status_inconsistencies", context["tables"])
        self.assertIn("commercial_pressure_summary", context["tables"])
        self.assertIn("driver_outliers", context["table_pagination"])
        self.assertEqual(context["table_pagination"]["driver_outliers"]["current_page"], 1)
        self.assertIn("cards", context["explanations"])
        self.assertIn("charts", context["explanations"])
        self.assertIn("tables", context["explanations"])
        self.assertIn("scores", context["explanations"])

        json.dumps(context)

    def test_cards_are_calculated_from_filtered_records(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            invoice_issue_date=date(2026, 5, 1),
            invoice_value=Decimal("100.00"),
            delivery_status="ENTREGUE",
            cargo_status="Ativa",
            operational_lead_time_hours=Decimal("10.00"),
            carrier_lead_time_hours=Decimal("5.00"),
            is_operational_late=True,
            is_carrier_late=False,
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            invoice_issue_date=date(2026, 5, 2),
            invoice_value=Decimal("50.50"),
            delivery_status="PENDENTE",
            operational_lead_time_hours=Decimal("20.00"),
            carrier_lead_time_hours=Decimal("7.00"),
            is_operational_late=False,
            is_carrier_late=True,
        )

        context = get_dashboard_context(DashboardFilters())
        cards = context["cards"]

        self.assertEqual(cards["total_records"], 2)
        self.assertEqual(cards["delivered_records"], 1)
        self.assertEqual(cards["pending_records"], 1)
        self.assertEqual(cards["total_invoice_value"], "150.50")
        self.assertEqual(cards["average_operational_lead_time_hours"], "15.00")
        self.assertEqual(cards["average_carrier_lead_time_hours"], "6.00")
        self.assertEqual(cards["operational_late_records"], 1)
        self.assertEqual(cards["carrier_late_records"], 1)
        self.assertEqual(cards["operational_late_percentage"], "50.00")
        self.assertEqual(cards["carrier_late_percentage"], "50.00")
        self.assertEqual(cards["operational_sla_rate"], "50.00")
        self.assertEqual(cards["carrier_sla_rate"], "50.00")
        self.assertEqual(cards["operational_lead_time_p90_hours"], "20.00")
        self.assertEqual(cards["carrier_lead_time_p90_hours"], "7.00")
        self.assertEqual(cards["delayed_invoice_value"], "150.50")
        self.assertEqual(cards["top_critical_route"]["route"], "RT030")
        self.assertEqual(cards["top_critical_route"]["criticality_score"], "70.00")
        self.assertEqual(cards["top_critical_route"]["total_records"], 2)
        self.assertEqual(cards["top_critical_route"]["delayed_percentage"], "100.00")
        self.assertEqual(cards["status_inconsistency_count"], 1)
        self.assertEqual(cards["status_inconsistency_percentage"], "50.00")

    def test_percentile_cards_handle_small_and_empty_sets(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            operational_lead_time_hours=Decimal("12.00"),
            carrier_lead_time_hours=None,
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            operational_lead_time_hours=Decimal("36.00"),
            carrier_lead_time_hours=None,
        )

        cards = get_dashboard_context(DashboardFilters())["cards"]

        self.assertEqual(cards["operational_lead_time_p90_hours"], "36.00")
        self.assertEqual(cards["carrier_lead_time_p90_hours"], "0.00")

    def test_delayed_invoice_value_counts_operational_or_carrier_late_records_once(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            invoice_value=Decimal("100.00"),
            is_operational_late=True,
            is_carrier_late=False,
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            invoice_value=Decimal("75.50"),
            is_operational_late=False,
            is_carrier_late=True,
        )
        self._create_record(
            batch=batch,
            row_number=3,
            invoice_number="1003",
            invoice_value=Decimal("30.00"),
            is_operational_late=True,
            is_carrier_late=True,
        )
        self._create_record(
            batch=batch,
            row_number=4,
            invoice_number="1004",
            invoice_value=Decimal("20.00"),
            is_operational_late=False,
            is_carrier_late=False,
        )

        cards = get_dashboard_context(DashboardFilters())["cards"]

        self.assertEqual(cards["delayed_invoice_value"], "205.50")

    def test_criticality_score_considers_delay_severity(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            route="RT_SEVERE",
            invoice_value=Decimal("100.00"),
            operational_lead_time_hours=Decimal("148.00"),
            carrier_lead_time_hours=Decimal("24.00"),
            is_operational_late=True,
            is_carrier_late=False,
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            route="RT_MILD",
            invoice_value=Decimal("100.00"),
            operational_lead_time_hours=Decimal("49.00"),
            carrier_lead_time_hours=Decimal("24.00"),
            is_operational_late=True,
            is_carrier_late=False,
        )

        routes = get_dashboard_context(DashboardFilters())["tables"]["critical_routes"]

        self.assertEqual(routes[0]["route"], "RT_SEVERE")
        self.assertEqual(routes[0]["delay_severity_hours"], "100.00")
        self.assertEqual(routes[1]["delay_severity_hours"], "1.00")

    def test_filters_are_applied_to_queryset(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            driver_name="BEATRIZ GOMES",
            route="RT030",
            business_unit="TABACO",
            region="SUL",
            frequency="DIARIA",
            delivery_status="ENTREGUE",
            cargo_status="Ativa",
            invoice_issue_date=date(2026, 5, 1),
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            driver_name="ANA SILVA",
            route="RT010",
            business_unit="ALIMENTOS",
            region="NORTE",
            frequency="SEMANAL",
            delivery_status="PENDENTE",
            cargo_status="Cancelada",
            invoice_issue_date=date(2026, 6, 1),
        )

        filters = DashboardFilters(
            date_start=date(2026, 5, 1),
            date_end=date(2026, 5, 31),
            driver_name="BEATRIZ GOMES",
            route="RT030",
            business_unit="TABACO",
            region="SUL",
            frequency="DIARIA",
            delivery_status="ENTREGUE",
            cargo_status="Ativa",
        )
        context = get_dashboard_context(filters)

        self.assertEqual(context["cards"]["total_records"], 1)
        self.assertEqual(context["tables"]["driver_outliers"][0]["driver_name"], "BEATRIZ GOMES")
        self.assertEqual(context["tables"]["critical_routes"][0]["route"], "RT030")
        self.assertEqual(context["tables"]["critical_regions"][0]["region"], "SUL")
        self.assertEqual(context["tables"]["critical_frequencies"][0]["frequency"], "DIARIA")

    def test_multi_select_filters_are_applied_to_queryset(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            driver_name="BEATRIZ GOMES",
            route="RT030",
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            driver_name="ANA SILVA",
            route="RT010",
        )
        self._create_record(
            batch=batch,
            row_number=3,
            invoice_number="1003",
            driver_name="CARLOS LIMA",
            route="RT099",
        )
        query = QueryDict("", mutable=True)
        query.setlist("driver_name", ["BEATRIZ GOMES", "ANA SILVA"])
        query.setlist("route", ["RT030", "RT010"])

        filters = DashboardFilters.from_querydict(query)
        context = get_dashboard_context(filters, query)

        self.assertEqual(filters.driver_name, ["BEATRIZ GOMES", "ANA SILVA"])
        self.assertEqual(filters.route, ["RT030", "RT010"])
        self.assertEqual(context["cards"]["total_records"], 2)
        self.assertEqual(
            {row["driver_name"] for row in context["tables"]["driver_outliers"]},
            {"ANA SILVA", "BEATRIZ GOMES"},
        )

    def test_table_pagination_preserves_filter_querystring(self):
        batch = self._create_batch()
        for index, driver_name in enumerate(
            [
                "ANA SILVA",
                "BEATRIZ GOMES",
                "CARLOS LIMA",
                "DANIEL ROCHA",
                "ELISA MOURA",
                "FERNANDO COSTA",
            ],
            start=1,
        ):
            self._create_record(
                batch=batch,
                row_number=index,
                invoice_number=str(1000 + index),
                driver_name=driver_name,
                route="RT030" if index % 2 else "RT010",
            )
        query = QueryDict("", mutable=True)
        query.setlist("route", ["RT030", "RT010"])
        query["driver_outliers_page"] = "2"

        context = get_dashboard_context(DashboardFilters.from_querydict(query), query)
        pagination = context["table_pagination"]["driver_outliers"]

        self.assertEqual(pagination["current_page"], 2)
        self.assertEqual(pagination["total_items"], 6)
        self.assertEqual(len(context["tables"]["driver_outliers"]), 1)
        self.assertIn("route=RT030", pagination["previous_url"])
        self.assertIn("route=RT010", pagination["previous_url"])
        self.assertIn("driver_outliers_page=1", pagination["previous_url"])

    def test_chart_contracts_have_stable_json_shape(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            driver_name="BEATRIZ GOMES",
            route="RT030",
            region="SUL",
            frequency="DIARIA",
            invoice_issue_date=date(2026, 5, 1),
            delivery_status="ENTREGUE",
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            driver_name="BEATRIZ GOMES",
            route="RT030",
            region="SUL",
            frequency="DIARIA",
            invoice_issue_date=date(2026, 5, 2),
            delivery_status="PENDENTE",
        )

        charts = get_dashboard_context(DashboardFilters())["charts"]

        self.assertEqual(charts["records_by_day"]["id"], "records_by_day")
        self.assertEqual(charts["records_by_day"]["type"], "bar")
        self.assertEqual(charts["records_by_day"]["data"]["labels"], ["2026-05-01", "2026-05-02"])
        self.assertEqual(charts["driver_efficiency_scatter"]["id"], "driver_efficiency_scatter")
        self.assertEqual(charts["driver_efficiency_scatter"]["type"], "bubble")
        self.assertEqual(charts["driver_efficiency_scatter"]["data"]["datasets"][0]["data"][0]["driver_name"], "BEATRIZ GOMES")
        self.assertEqual(charts["critical_routes_ranking"]["id"], "critical_routes_ranking")
        self.assertEqual(charts["critical_routes_ranking"]["options"]["indexAxis"], "y")
        self.assertEqual(charts["weekday_bottleneck"]["id"], "weekday_bottleneck")
        self.assertEqual(len(charts["weekday_bottleneck"]["data"]["datasets"]), 2)
        self.assertEqual(charts["delay_pareto"]["id"], "delay_pareto")
        self.assertEqual(charts["delay_pareto"]["data"]["datasets"][1]["type"], "line")
        self.assertEqual(charts["lead_time_distribution"]["id"], "lead_time_distribution")
        self.assertEqual(len(charts["lead_time_distribution"]["data"]["datasets"]), 2)
        self.assertEqual(charts["region_lead_time_comparison"]["id"], "region_lead_time_comparison")
        self.assertEqual(charts["region_lead_time_comparison"]["data"]["labels"], ["SUL"])
        self.assertEqual(charts["frequency_lead_time_comparison"]["id"], "frequency_lead_time_comparison")
        self.assertEqual(charts["frequency_lead_time_comparison"]["data"]["labels"], ["DIARIA"])

        json.dumps(charts)

    def test_commercial_pressure_contracts_compare_billing_and_delivery(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            invoice_issue_date=date(2026, 5, 26),
            customer_delivery_date=date(2026, 5, 27),
            operational_lead_time_hours=Decimal("12.00"),
            carrier_lead_time_hours=Decimal("6.00"),
        )
        self._create_record(
            batch=batch,
            row_number=2,
            invoice_number="1002",
            invoice_issue_date=date(2026, 5, 27),
            customer_delivery_date=date(2026, 5, 29),
            operational_lead_time_hours=Decimal("80.00"),
            carrier_lead_time_hours=Decimal("30.00"),
            is_operational_late=True,
            is_carrier_late=True,
        )
        self._create_record(
            batch=batch,
            row_number=3,
            invoice_number="1003",
            invoice_issue_date=date(2026, 5, 27),
            customer_delivery_date=date(2026, 5, 29),
            operational_lead_time_hours=Decimal("90.00"),
            carrier_lead_time_hours=Decimal("40.00"),
            is_operational_late=True,
            is_carrier_late=True,
        )

        context = get_dashboard_context(DashboardFilters())

        self.assertEqual(context["cards"]["peak_billing_day"]["date"], "2026-05-27")
        self.assertEqual(context["cards"]["peak_billing_day"]["records"], 2)
        self.assertEqual(context["cards"]["last_3_business_days_records"], 2)
        self.assertEqual(context["cards"]["last_3_business_days_percentage"], "66.67")
        self.assertEqual(
            context["charts"]["billing_vs_delivery_by_day"]["data"]["labels"],
            ["2026-05-26", "2026-05-27", "2026-05-29"],
        )
        self.assertEqual(
            context["charts"]["billing_vs_delivery_by_day"]["data"]["datasets"][0]["data"],
            [1, 2, 0],
        )
        self.assertEqual(
            context["charts"]["billing_vs_delivery_by_day"]["data"]["datasets"][1]["data"],
            [0, 1, 2],
        )
        self.assertEqual(context["tables"]["commercial_pressure_summary"][0]["period"], "Período normal")
        self.assertEqual(context["tables"]["commercial_pressure_summary"][0]["records"], 1)
        self.assertEqual(context["tables"]["commercial_pressure_summary"][1]["period"], "Últimos 3 dias úteis")
        self.assertEqual(context["tables"]["commercial_pressure_summary"][1]["records"], 2)
        self.assertEqual(
            context["tables"]["commercial_pressure_summary"][1]["operational_late_percentage"],
            "100.00",
        )

    def test_new_dashboard_tables_are_available(self):
        batch = self._create_batch()
        self._create_record(
            batch=batch,
            row_number=1,
            invoice_number="1001",
            driver_name="BEATRIZ GOMES",
            route="RT030",
            city="PONTA GROSSA",
            region="SUL",
            frequency="DIARIA",
            invoice_value=Decimal("100.00"),
            operational_lead_time_hours=Decimal("120.00"),
            carrier_lead_time_hours=Decimal("72.00"),
            is_operational_late=True,
            cargo_status="Ativa",
            delivery_status="ENTREGUE",
        )

        tables = get_dashboard_context(DashboardFilters())["tables"]

        self.assertEqual(tables["driver_outliers"][0]["driver_name"], "BEATRIZ GOMES")
        self.assertEqual(tables["critical_routes"][0]["route"], "RT030")
        self.assertEqual(tables["critical_routes"][0]["served_cities"], 1)
        self.assertEqual(tables["critical_cities"][0]["city"], "PONTA GROSSA")
        self.assertEqual(tables["critical_regions"][0]["region"], "SUL")
        self.assertEqual(tables["critical_frequencies"][0]["frequency"], "DIARIA")
        self.assertEqual(tables["invoice_outliers"][0]["invoice_number"], "1001")
        self.assertEqual(tables["status_inconsistencies"][0]["invoice_number"], "1001")

        json.dumps(tables)

    def test_last_successful_import_ignores_failed_batches(self):
        self._create_batch(status=ImportBatch.Status.ERROR, file_hash="error-hash")
        success_batch = self._create_batch(status=ImportBatch.Status.SUCCESS, file_hash="success-hash")

        metadata = get_dashboard_context(DashboardFilters())["metadata"]

        self.assertEqual(metadata["last_successful_import"]["id"], success_batch.id)
        self.assertEqual(metadata["last_successful_import"]["status"], ImportBatch.Status.SUCCESS)

    def test_invalid_date_filter_is_ignored_safely(self):
        filters = DashboardFilters.from_querydict({"date_start": "invalid-date"})

        self.assertIsNone(filters.date_start)
        self.assertEqual(filters.errors, ["date_start must use YYYY-MM-DD."])
        self.assertEqual(get_dashboard_context(filters)["cards"]["total_records"], 0)

    def _create_batch(self, status=ImportBatch.Status.SUCCESS, file_hash="file-hash"):
        now = timezone.now()
        return ImportBatch.objects.create(
            source_mode=ImportBatch.SourceMode.LOCAL,
            file_name=f"{file_hash}.xlsx",
            file_path="data/test.xlsx",
            file_hash=file_hash,
            sheet_name="COM 001",
            status=status,
            started_at=now,
            finished_at=now if status == ImportBatch.Status.SUCCESS else None,
            total_rows=1,
            valid_rows=1 if status == ImportBatch.Status.SUCCESS else 0,
            invalid_rows=0,
        )

    def _create_record(self, batch, row_number, invoice_number, **overrides):
        defaults = {
            "row_hash": f"row-hash-{row_number}",
            "business_unit": "TABACO",
            "map_date": date(2026, 5, 1),
            "map_number": "470939",
            "owner": "948 TAINARA CAMARGO MONTE",
            "route": "RT030",
            "delivery_points": 18,
            "weight": Decimal("151.670"),
            "volume": Decimal("0.640"),
            "map_value": Decimal("375.10"),
            "vehicle_plate": "AZP8G94",
            "driver_code": "395",
            "driver_name": "BEATRIZ GOMES",
            "checker_code": "",
            "checker_name": "",
            "load_status_description": "CARREGADO EM",
            "load_date": date(2026, 5, 5),
            "invoice_series": "11",
            "invoice_issue_date": date(2026, 5, 4),
            "bordero_number": "807679",
            "customer_code": "6248888",
            "customer_name": "CLIENTE TESTE LTDA",
            "city": "PONTA GROSSA",
            "region": "",
            "frequency": "",
            "invoice_value": Decimal("1674.40"),
            "invoice_status": "AZP8G94",
            "cargo_status": "Ativa",
            "auxiliary_date": None,
            "delivery_status": "ENTREGUE",
            "customer_delivery_date": date(2026, 5, 5),
            "customer_delivery_time": time(16, 12),
            "customer_delivery_datetime": timezone.now(),
            "exported_to": "UMOV",
            "notes": "",
            "seller_code": "681",
            "team_code": "1",
            "operational_lead_time_hours": Decimal("10.00"),
            "carrier_lead_time_hours": Decimal("5.00"),
            "is_operational_late": False,
            "is_carrier_late": False,
            "business_days_count": 1,
        }
        defaults.update(overrides)

        return LeadTimeRecord.objects.create(
            import_batch=batch,
            row_number=row_number,
            invoice_number=invoice_number,
            **defaults,
        )
