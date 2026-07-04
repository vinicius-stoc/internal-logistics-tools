from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.db.models import Avg, Count, Q, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from imports.models import LeadTimeRecord

from .dashboard_filters import DashboardFilters
from .dashboard_queries import get_dashboard_context


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DELIVERED_STATUS_LOOKUP = Q(delivery_status__icontains="entreg")

TABLE_EXPORTS = {
    "commercial_pressure_summary": {
        "sheet_title": "Pressao comercial",
        "file_suffix": "pressao_comercial",
        "columns": [
            ("Periodo", "period", None),
            ("Dias", "issue_days", "integer"),
            ("Registros", "records", "integer"),
            ("Participacao", "records_share", "decimal"),
            ("Media diaria", "daily_average_records", "decimal"),
            ("Valor NF", "total_invoice_value", "currency"),
            ("Atraso operacional", "operational_late_percentage", "decimal"),
            ("Atraso transportadora", "carrier_late_percentage", "decimal"),
            ("LT operacional medio (h)", "average_operational_lead_time_hours", "decimal"),
            ("P90 operacional (h)", "operational_lead_time_p90_hours", "decimal"),
            ("Severidade atraso (h)", "delay_severity_hours", "decimal"),
        ],
    },
    "driver_outliers": {
        "sheet_title": "Motoristas fora da curva",
        "file_suffix": "motoristas_fora_da_curva",
        "columns": [
            ("Motorista", "driver_name", None),
            ("Registros", "total_records", "integer"),
            ("Entregues", "delivered_records", "integer"),
            ("Atrasados", "delayed_records", "integer"),
            ("Valor NF", "total_invoice_value", "currency"),
            ("Severidade atraso (h)", "delay_severity_hours", "decimal"),
            ("LT operacional medio (h)", "average_operational_lead_time_hours", "decimal"),
            ("P90 operacional (h)", "operational_lead_time_p90_hours", "decimal"),
            ("LT transportadora medio (h)", "average_carrier_lead_time_hours", "decimal"),
            ("Atraso operacional", "operational_late_percentage", "decimal"),
            ("Atraso transportadora", "carrier_late_percentage", "decimal"),
            ("Score", "criticality_score", "decimal"),
        ],
    },
    "invoice_outliers": {
        "sheet_title": "Notas fora da curva",
        "file_suffix": "notas_fora_da_curva",
        "columns": [
            ("NF", "invoice_number", None),
            ("Serie", "invoice_series", None),
            ("Motorista", "driver_name", None),
            ("Pauta/Rota", "route", None),
            ("Cidade", "city", None),
            ("Valor NF", "invoice_value", "currency"),
            ("Emissao", "invoice_issue_date", None),
            ("Entrega", "customer_delivery_date", None),
            ("LT operacional (h)", "operational_lead_time_hours", "decimal"),
            ("LT transportadora (h)", "carrier_lead_time_hours", "decimal"),
            ("Status carga", "cargo_status", None),
            ("Status entrega", "delivery_status", None),
        ],
    },
    "status_inconsistencies": {
        "sheet_title": "Divergencias de status",
        "file_suffix": "divergencias_status",
        "columns": [
            ("NF", "invoice_number", None),
            ("Serie", "invoice_series", None),
            ("Motorista", "driver_name", None),
            ("Pauta/Rota", "route", None),
            ("Cidade", "city", None),
            ("Emissao", "invoice_issue_date", None),
            ("Entrega", "customer_delivery_date", None),
            ("Status carga", "cargo_status", None),
            ("Status entrega", "delivery_status", None),
            ("LT operacional (h)", "operational_lead_time_hours", "decimal"),
            ("LT transportadora (h)", "carrier_lead_time_hours", "decimal"),
        ],
    },
    "critical_routes": {
        "sheet_title": "Rotas criticas",
        "file_suffix": "rotas_criticas",
        "columns": [
            ("Pauta/Rota", "route", None),
            ("Registros", "total_records", "integer"),
            ("Entregues", "delivered_records", "integer"),
            ("Atrasados", "delayed_records", "integer"),
            ("Cidades atendidas", "served_cities", "integer"),
            ("Valor NF", "total_invoice_value", "currency"),
            ("Severidade atraso (h)", "delay_severity_hours", "decimal"),
            ("LT operacional medio (h)", "average_operational_lead_time_hours", "decimal"),
            ("P90 operacional (h)", "operational_lead_time_p90_hours", "decimal"),
            ("Atraso operacional", "operational_late_percentage", "decimal"),
            ("Atraso transportadora", "carrier_late_percentage", "decimal"),
            ("Score", "criticality_score", "decimal"),
        ],
    },
    "critical_cities": {
        "sheet_title": "Cidades criticas",
        "file_suffix": "cidades_criticas",
        "columns": [
            ("Cidade", "city", None),
            ("Registros", "total_records", "integer"),
            ("Rotas", "routes", "integer"),
            ("Atrasados", "delayed_records", "integer"),
            ("Valor NF", "total_invoice_value", "currency"),
            ("LT operacional medio (h)", "average_operational_lead_time_hours", "decimal"),
            ("LT transportadora medio (h)", "average_carrier_lead_time_hours", "decimal"),
            ("Severidade atraso (h)", "delay_severity_hours", "decimal"),
            ("Atraso operacional", "operational_late_percentage", "decimal"),
            ("Atraso transportadora", "carrier_late_percentage", "decimal"),
            ("Score", "criticality_score", "decimal"),
        ],
    },
    "critical_regions": {
        "sheet_title": "Regioes criticas",
        "file_suffix": "regioes_criticas",
        "columns": [
            ("Regiao", "region", None),
            ("Registros", "total_records", "integer"),
            ("Atrasados", "delayed_records", "integer"),
            ("Valor NF", "total_invoice_value", "currency"),
            ("LT operacional medio (h)", "average_operational_lead_time_hours", "decimal"),
            ("LT transportadora medio (h)", "average_carrier_lead_time_hours", "decimal"),
            ("Severidade atraso (h)", "delay_severity_hours", "decimal"),
            ("Atraso operacional", "operational_late_percentage", "decimal"),
            ("Atraso transportadora", "carrier_late_percentage", "decimal"),
            ("Score", "criticality_score", "decimal"),
        ],
    },
    "critical_frequencies": {
        "sheet_title": "Frequencias criticas",
        "file_suffix": "frequencias_criticas",
        "columns": [
            ("Frequencia", "frequency", None),
            ("Registros", "total_records", "integer"),
            ("Atrasados", "delayed_records", "integer"),
            ("Valor NF", "total_invoice_value", "currency"),
            ("LT operacional medio (h)", "average_operational_lead_time_hours", "decimal"),
            ("LT transportadora medio (h)", "average_carrier_lead_time_hours", "decimal"),
            ("Severidade atraso (h)", "delay_severity_hours", "decimal"),
            ("Atraso operacional", "operational_late_percentage", "decimal"),
            ("Atraso transportadora", "carrier_late_percentage", "decimal"),
            ("Score", "criticality_score", "decimal"),
        ],
    },
}


@dataclass(frozen=True)
class DashboardExportResult:
    file_name: str
    content: bytes
    content_type: str = XLSX_CONTENT_TYPE


class DashboardExportError(Exception):
    pass


def build_dashboard_export(filters: DashboardFilters, table_key=None) -> DashboardExportResult:
    try:
        if table_key:
            return _build_table_export(filters, table_key)

        queryset = _get_filtered_queryset(filters)
        workbook = Workbook()

        data_sheet = workbook.active
        data_sheet.title = "Dados filtrados"
        _fill_data_sheet(data_sheet, queryset)
        _fill_summary_sheet(workbook, queryset, "Resumo motorista", "driver_name", "Motorista")
        _fill_summary_sheet(workbook, queryset, "Resumo pauta", "route", "Pauta")
        _fill_summary_sheet(workbook, queryset, "Resumo região", "region", "Região")
        _fill_summary_sheet(workbook, queryset, "Resumo frequência", "frequency", "Frequência")

        output = BytesIO()
        workbook.save(output)

        return DashboardExportResult(
            file_name=_make_file_name(),
            content=output.getvalue(),
        )
    except Exception as exc:
        raise DashboardExportError("Não foi possível gerar a exportação Excel.") from exc


def _build_table_export(filters, table_key):
    table_config = TABLE_EXPORTS.get(table_key)
    if not table_config:
        raise DashboardExportError("Tabela inválida para exportação Excel.")

    context = get_dashboard_context(filters, paginate_tables=False)
    rows = context["tables"].get(table_key, [])
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = table_config["sheet_title"]
    _fill_table_sheet(worksheet, rows, table_config["columns"])

    output = BytesIO()
    workbook.save(output)

    return DashboardExportResult(
        file_name=_make_file_name(table_config["file_suffix"]),
        content=output.getvalue(),
    )


def _fill_table_sheet(worksheet, rows, columns):
    worksheet.append([column[0] for column in columns])

    for row in rows:
        worksheet.append(
            [
                _format_table_value(row.get(field_name), value_type)
                for _, field_name, value_type in columns
            ]
        )

    _format_worksheet(worksheet, columns)


def _get_filtered_queryset(filters):
    queryset = LeadTimeRecord.objects.all().only(
        "business_unit",
        "invoice_issue_date",
        "customer_delivery_date",
        "driver_name",
        "route",
        "city",
        "region",
        "frequency",
        "customer_name",
        "invoice_number",
        "invoice_series",
        "invoice_value",
        "cargo_status",
        "delivery_status",
        "operational_lead_time_hours",
        "carrier_lead_time_hours",
        "is_operational_late",
        "is_carrier_late",
        "vehicle_plate",
        "map_number",
        "delivery_points",
        "weight",
        "volume",
    )
    return filters.apply_to_queryset(queryset).order_by(
        "invoice_issue_date",
        "driver_name",
        "route",
        "invoice_number",
        "id",
    )


def _fill_data_sheet(worksheet, queryset):
    columns = [
        ("Unidade de negócio", "business_unit", None),
        ("Data emissão NF", "invoice_issue_date", "date"),
        ("Data entrega cliente", "customer_delivery_date", "date"),
        ("Motorista", "driver_name", None),
        ("Pauta", "route", None),
        ("Cidade", "city", None),
        ("Região", "region", None),
        ("Frequência", "frequency", None),
        ("Cliente", "customer_name", None),
        ("NF", "invoice_number", None),
        ("Série", "invoice_series", None),
        ("Valor NF", "invoice_value", "currency"),
        ("Status carga", "cargo_status", None),
        ("Status entrega", "delivery_status", None),
        ("Lead time operacional (h)", "operational_lead_time_hours", "decimal"),
        ("Lead time transportadora (h)", "carrier_lead_time_hours", "decimal"),
        ("Atraso operacional", "is_operational_late", "boolean"),
        ("Atraso transportadora", "is_carrier_late", "boolean"),
        ("Placa", "vehicle_plate", None),
        ("Mapa", "map_number", None),
        ("Pontos de entrega", "delivery_points", "integer"),
        ("Peso", "weight", "decimal"),
        ("Volume", "volume", "decimal"),
    ]

    worksheet.append([column[0] for column in columns])

    for record in queryset.iterator(chunk_size=1000):
        worksheet.append(
            [
                _format_cell_value(getattr(record, field_name), value_type)
                for _, field_name, value_type in columns
            ]
        )

    _format_worksheet(worksheet, columns)


def _fill_summary_sheet(workbook, queryset, title, field_name, label):
    worksheet = workbook.create_sheet(title=title)
    columns = [
        (label, field_name, None),
        ("Registros", "total_records", "integer"),
        ("Entregues", "delivered_records", "integer"),
        ("Pendentes", "pending_records", "integer"),
        ("Valor NF", "total_invoice_value", "currency"),
        ("Lead time operacional médio (h)", "average_operational_lead_time_hours", "decimal"),
        ("Lead time transportadora médio (h)", "average_carrier_lead_time_hours", "decimal"),
        ("Atraso operacional", "operational_late_records", "integer"),
        ("Atraso transportadora", "carrier_late_records", "integer"),
    ]
    worksheet.append([column[0] for column in columns])

    rows = (
        queryset.values(field_name)
        .annotate(
            total_records=Count("id"),
            delivered_records=Count("id", filter=DELIVERED_STATUS_LOOKUP),
            operational_late_records=Count("id", filter=Q(is_operational_late=True)),
            carrier_late_records=Count("id", filter=Q(is_carrier_late=True)),
            total_invoice_value=Sum("invoice_value"),
            average_operational_lead_time_hours=Avg("operational_lead_time_hours"),
            average_carrier_lead_time_hours=Avg("carrier_lead_time_hours"),
        )
        .order_by("-total_records", field_name)
    )

    for row in rows:
        pending_records = row["total_records"] - row["delivered_records"]
        worksheet.append(
            [
                row[field_name] or f"Sem {label.lower()}",
                row["total_records"],
                row["delivered_records"],
                pending_records,
                _decimal_to_number(row["total_invoice_value"]),
                _decimal_to_number(row["average_operational_lead_time_hours"]),
                _decimal_to_number(row["average_carrier_lead_time_hours"]),
                row["operational_late_records"],
                row["carrier_late_records"],
            ]
        )

    _format_worksheet(worksheet, columns)


def _format_cell_value(value, value_type):
    if value is None:
        return ""
    if value_type == "boolean":
        return "Sim" if value else "Não"
    if value_type in {"currency", "decimal"}:
        return _decimal_to_number(value)
    if isinstance(value, (date, datetime)):
        return value
    return value


def _format_table_value(value, value_type):
    if value is None:
        return ""
    if value_type in {"currency", "decimal"}:
        return _decimal_to_number(Decimal(str(value)))
    if value_type == "integer":
        return int(value)
    return value


def _decimal_to_number(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    return value


def _format_worksheet(worksheet, columns):
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for index, (_, _, value_type) in enumerate(columns, start=1):
        column_letter = get_column_letter(index)
        if value_type == "date":
            number_format = "dd/mm/yyyy"
        elif value_type == "currency":
            number_format = 'R$ #,##0.00'
        elif value_type == "decimal":
            number_format = "0.00"
        elif value_type == "integer":
            number_format = "0"
        else:
            number_format = None

        max_length = 0
        for cell in worksheet[column_letter]:
            if number_format and cell.row > 1 and cell.value != "":
                cell.number_format = number_format
            max_length = max(max_length, len(str(cell.value or "")))

        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _make_file_name(suffix=None):
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    suffix_part = f"_{suffix}" if suffix else ""
    return f"dashboard_lead_time{suffix_part}_{timestamp}.xlsx"
